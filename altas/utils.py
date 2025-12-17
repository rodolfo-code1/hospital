# hospital/altas/utils.py
import os
from django.http import HttpResponse
from django.conf import settings
# Librerías para generación de PDF (ReportLab)
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
# Librerías para Excel (OpenPyXL)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
# Manejo de zonas horarias (Crítico para que la hora impresa sea la local, no UTC)
from django.utils import timezone 

def generar_certificado_pdf(alta):
    """
    Genera el Certificado de Alta Hospitalaria en formato PDF.
    
    Este documento sirve como comprobante oficial de egreso para el paciente.
    El contenido es dinámico: se adapta dependiendo de si el alta incluye a la madre,
    al recién nacido, o a ambos.
    
    Args:
        alta (Alta): Objeto del modelo Alta con los datos a imprimir.
        
    Returns:
        HttpResponse: Respuesta HTTP con el archivo PDF adjunto (application/pdf).
    """
    response = HttpResponse(content_type='application/pdf')
    
    # 1. GENERACIÓN DE NOMBRE DE ARCHIVO DINÁMICO
    # El nombre del archivo ayuda a identificar rápidamente de quién es el certificado.
    if alta.madre and alta.recien_nacido:
        sufijo = f"{alta.madre.rut}_RN"
    elif alta.madre:
        sufijo = f"{alta.madre.rut}"
    elif alta.recien_nacido:
        sufijo = f"RN_{alta.recien_nacido.codigo_unico}"
    else:
        sufijo = f"Alta_{alta.id}"
        
    # Usamos la hora actual LOCAL para el nombre del archivo (evita confusión UTC)
    fecha_hoy = timezone.localtime(timezone.now())
    filename = f'certificado_alta_{sufijo}_{fecha_hoy.strftime("%Y%m%d")}.pdf'
    
    # Content-Disposition: attachment fuerza la descarga en el navegador
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # 2. CONFIGURACIÓN DEL DOCUMENTO
    doc = SimpleDocTemplate(response, pagesize=letter)
    story = [] # La "historia" es la lista de elementos (párrafos, tablas) que se pintarán en orden
    styles = getSampleStyleSheet()
    
    # Definición de Estilos Personalizados (Azul corporativo)
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a365d'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    subtitulo_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2c5282'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    texto_normal = styles['Normal']
    
    # --- ENCABEZADO DEL DOCUMENTO ---
    story.append(Paragraph("CERTIFICADO DE ALTA HOSPITALARIA", titulo_style))
    story.append(Spacer(1, 0.2 * inch))
    
    # Formateo de fecha legible para el usuario
    if alta.fecha_alta:
        fecha_alta_local = timezone.localtime(alta.fecha_alta)
        fecha_texto = fecha_alta_local.strftime('%d/%m/%Y a las %H:%M')
    else:
        fecha_texto = "Fecha pendiente"

    intro = f"El Hospital Clínico Herminda Martín certifica que el proceso de alta médica y administrativa ha concluido exitosamente con fecha {fecha_texto}."
    story.append(Paragraph(intro, texto_normal))
    story.append(Spacer(1, 0.3 * inch))
    
    # --- SECCIÓN 1: DATOS DE LA MADRE (Condicional) ---
    if alta.madre:
        story.append(Paragraph("DATOS DE LA MADRE", subtitulo_style))
        
        datos_madre = [
            ['Nombre Completo:', alta.madre.nombre],
            ['RUT:', alta.madre.rut],
            ['Edad:', f'{alta.madre.edad} años'],
            ['Previsión:', alta.madre.get_prevision_display()],
            ['Dirección:', f"{alta.madre.direccion}, {alta.madre.comuna}"],
        ]
        
        # Tabla con estilos para organizar la información clave
        tabla_madre = Table(datos_madre, colWidths=[2*inch, 4*inch])
        tabla_madre.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e2e8f0')), # Fondo gris azulado para etiquetas
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(tabla_madre)
        story.append(Spacer(1, 0.2 * inch))

    # --- SECCIÓN 2: DATOS DEL RECIÉN NACIDO (Condicional) ---
    if alta.recien_nacido:
        story.append(Paragraph("DATOS DEL RECIÉN NACIDO", subtitulo_style))
        
        datos_rn = [
            ['Código ID:', alta.recien_nacido.codigo_unico],
            ['Sexo:', alta.recien_nacido.get_sexo_display()],
            ['Peso al Nacer:', f'{alta.recien_nacido.peso} kg'],
            ['Talla:', f'{alta.recien_nacido.talla} cm'],
            ['APGAR (1/5):', f'{alta.recien_nacido.apgar_1_min} / {alta.recien_nacido.apgar_5_min}'],
        ]
        
        tabla_rn = Table(datos_rn, colWidths=[2*inch, 4*inch])
        tabla_rn.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e2e8f0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(tabla_rn)
        story.append(Spacer(1, 0.2 * inch))

    # --- SECCIÓN 3: ANTECEDENTES DEL PARTO (Condicional) ---
    if alta.parto:
        story.append(Paragraph("ANTECEDENTES DEL PARTO", subtitulo_style))
        
        fecha_parto = timezone.localtime(alta.parto.fecha_hora_inicio).strftime('%d/%m/%Y')

        datos_parto = [
            ['Tipo:', alta.parto.get_tipo_display()],
            ['Fecha:', fecha_parto],
            ['Médico:', alta.parto.medico_responsable],
            ['Matrona:', alta.parto.matrona_responsable],
        ]
        
        tabla_parto = Table(datos_parto, colWidths=[2*inch, 4*inch])
        tabla_parto.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f7fafc')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ]))
        story.append(tabla_parto)
        story.append(Spacer(1, 0.2 * inch))

    # --- SECCIÓN 4: ENTREGA DE ANTICONCEPTIVOS (MAC) ---
    # Solo se muestra si efectivamente se entregó un método
    if alta.se_entrego_anticonceptivo:
        story.append(Paragraph("PLANIFICACIÓN FAMILIAR / MAC", subtitulo_style))
        datos_mac = [
            ['Método Entregado:', alta.get_metodo_anticonceptivo_display()],
            ['Estado:', 'Suministrado al alta'],
        ]
        tabla_mac = Table(datos_mac, colWidths=[2*inch, 4*inch])
        tabla_mac.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6fffa')), # Fondo verdoso suave
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(tabla_mac)
        story.append(Spacer(1, 0.2 * inch))

    # --- SECCIÓN 5: FIRMAS Y RESPONSABLES ---
    story.append(Paragraph("RESPONSABLES DEL ALTA", subtitulo_style))
    
    # Formateo de fechas de firmas (si existen)
    fecha_clinica_str = '-'
    if alta.fecha_confirmacion_clinica:
        fecha_clinica_str = timezone.localtime(alta.fecha_confirmacion_clinica).strftime('%d/%m/%Y %H:%M')

    fecha_admin_str = '-'
    if alta.fecha_confirmacion_administrativa:
        fecha_admin_str = timezone.localtime(alta.fecha_confirmacion_administrativa).strftime('%d/%m/%Y %H:%M')

    datos_alta = [
        ['Alta Médica:', alta.medico_confirma or '-'],
        ['Fecha Clínica:', fecha_clinica_str],
        ['Cierre Administrativo:', alta.administrativo_confirma or '-'],
        ['Fecha Admin:', fecha_admin_str],
    ]
    
    tabla_alta = Table(datos_alta, colWidths=[2*inch, 4*inch])
    tabla_alta.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    story.append(tabla_alta)
    story.append(Spacer(1, 0.5 * inch))
    
    # --- OBSERVACIONES ---
    if alta.observaciones:
        story.append(Paragraph("OBSERVACIONES / INDICACIONES", subtitulo_style))
        # Reemplazamos saltos de línea normales por saltos HTML para que ReportLab los entienda
        obs_text = Paragraph(alta.observaciones.replace('\n', '<br/>'), texto_normal)
        story.append(obs_text)
    
    # --- PIE DE PÁGINA ---
    pie_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    fecha_emision = Paragraph(
        f"Documento generado el {fecha_hoy.strftime('%d/%m/%Y a las %H:%M')}",
        pie_style
    )
    story.append(Spacer(1, 0.5 * inch))
    story.append(fecha_emision)
    
    # Genera el PDF en el objeto response
    doc.build(story)
    return response


def exportar_altas_excel(altas):
    """
    Genera un reporte en formato Excel (.xlsx) con el listado de altas proporcionado.
    Útil para la jefatura administrativa y análisis de datos masivos.
    
    Args:
        altas (QuerySet): Conjunto de objetos Alta a exportar.
        
    Returns:
        HttpResponse: Archivo Excel para descargar.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro de Altas"
    
    # Estilos para el encabezado (Azul oscuro con texto blanco)
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = [
        'Folio', 'Tipo Alta', 'RUT Madre', 'Nombre Madre', 
        'Código RN', 'Sexo RN', 
        'Entrega MAC', 'Método MAC',
        'Médico Resp.', 'Admin Resp.', 'Fecha Alta', 'Estado'
    ]
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Escribir datos iterando sobre el QuerySet
    for row_num, alta in enumerate(altas, 2):
        # Determinar tipo de alta para estadística rápida
        tipo = "Conjunta"
        if not alta.madre: tipo = "Solo RN"
        if not alta.recien_nacido: tipo = "Solo Madre"
        
        # Manejo seguro de valores nulos (si no hay madre o RN, se pone "-")
        rut_madre = alta.madre.rut if alta.madre else "-"
        nom_madre = alta.madre.nombre if alta.madre else "-"
        cod_rn = alta.recien_nacido.codigo_unico if alta.recien_nacido else "-"
        sexo_rn = alta.recien_nacido.get_sexo_display() if alta.recien_nacido else "-"
        
        mac_entregado = "SÍ" if alta.se_entrego_anticonceptivo else "NO"
        mac_metodo = alta.get_metodo_anticonceptivo_display() if alta.se_entrego_anticonceptivo else "-"

        # Formateo de fecha para Excel (sin zona horaria para compatibilidad)
        fecha_alta_str = '-'
        if alta.fecha_alta:
            fecha_local = timezone.localtime(alta.fecha_alta)
            fecha_alta_str = fecha_local.strftime('%d/%m/%Y %H:%M')

        ws.cell(row=row_num, column=1, value=alta.id)
        ws.cell(row=row_num, column=2, value=tipo)
        ws.cell(row=row_num, column=3, value=rut_madre)
        ws.cell(row=row_num, column=4, value=nom_madre)
        ws.cell(row=row_num, column=5, value=cod_rn)
        ws.cell(row=row_num, column=6, value=sexo_rn)
        ws.cell(row=row_num, column=7, value=mac_entregado)
        ws.cell(row=row_num, column=8, value=mac_metodo)
        ws.cell(row=row_num, column=9, value=alta.medico_confirma or '-')
        ws.cell(row=row_num, column=10, value=alta.administrativo_confirma or '-')
        ws.cell(row=row_num, column=11, value=fecha_alta_str)
        ws.cell(row=row_num, column=12, value=alta.get_estado_display())
    
    # Ajuste automático del ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2
    
    # Configurar respuesta HTTP para descarga de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    fecha_nombre = timezone.localtime(timezone.now())
    filename = f'reporte_altas_{fecha_nombre.strftime("%Y%m%d_%H%M")}.xlsx'
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response