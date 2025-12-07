# reportes/excel_export.py
"""
Módulo para exportar reportes REM a formato Excel.
"""
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.utils import timezone
from partos.models import Parto, Aborto
from recien_nacidos.models import RecienNacido
from altas.models import Alta
from pacientes.models import Madre
from django.db.models import Count, Q


def crear_libro_excel_seccion_a(fecha_inicio, fecha_fin):
    """
    Exporta Sección A: Información General de Partos (REM)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sección A - Partos"
    
    # Estilos
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:H1')
    titulo = ws['A1']
    titulo.value = f"SECCIÓN A: INFORMACIÓN GENERAL DE PARTOS"
    titulo.font = Font(bold=True, size=14)
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:H2')
    subtitulo = ws['A2']
    subtitulo.value = f"Período: {fecha_inicio} a {fecha_fin}"
    subtitulo.font = Font(size=10, italic=True)
    subtitulo.alignment = Alignment(horizontal='center')
    
    # Obtener datos
    partos = Parto.objects.filter(fecha_hora_inicio__date__range=[fecha_inicio, fecha_fin])
    total_partos = partos.count()
    
    # Fila 4: Encabezados principales
    row = 4
    headers = ['Característica', 'Total', 'Menor 15 años', '15-19 años', '20-34 años', 'Mayor 35 años', 'RN <2.5kg', 'RN ≥2.5kg']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # 1. Total de Partos
    ws.cell(row=row, column=1).value = "Total de Partos"
    ws.cell(row=row, column=2).value = total_partos
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1
    
    # 2. Partos por Tipo
    tipos_partos = partos.values('tipo').annotate(count=Count('id')).order_by('tipo')
    for tipo_parto in tipos_partos:
        tipo = tipo_parto['tipo'] or 'Sin especificar'
        count = tipo_parto['count']
        
        # Contar por edad de madre
        partos_tipo = partos.filter(tipo=tipo_parto['tipo'] if tipo_parto['tipo'] else None)
        
        edad_menores_15 = partos_tipo.filter(madre__edad__lt=15).count()
        edad_15_19 = partos_tipo.filter(madre__edad__gte=15, madre__edad__lte=19).count()
        edad_20_34 = partos_tipo.filter(madre__edad__gte=20, madre__edad__lte=34).count()
        edad_mayor_35 = partos_tipo.filter(madre__edad__gt=35).count()
        
        # RN por peso
        rn_menor_2_5 = partos_tipo.filter(recien_nacidos__peso__lt=2.5).count()
        rn_mayor_2_5 = partos_tipo.filter(recien_nacidos__peso__gte=2.5).count()
        
        ws.cell(row=row, column=1).value = f"Parto {tipo}"
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = edad_menores_15
        ws.cell(row=row, column=4).value = edad_15_19
        ws.cell(row=row, column=5).value = edad_20_34
        ws.cell(row=row, column=6).value = edad_mayor_35
        ws.cell(row=row, column=7).value = rn_menor_2_5
        ws.cell(row=row, column=8).value = rn_mayor_2_5
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        row += 1
    
    # 3. Partos Prematuros
    row += 1
    ws.cell(row=row, column=1).value = "PARTOS PREMATUROS"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    prematuros_data = [
        ('<24 semanas', partos.filter(edad_gestacional_semanas__lt=24).count()),
        ('24-28 semanas', partos.filter(edad_gestacional_semanas__gte=24, edad_gestacional_semanas__lte=28).count()),
        ('29-32 semanas', partos.filter(edad_gestacional_semanas__gte=29, edad_gestacional_semanas__lte=32).count()),
        ('33-36 semanas', partos.filter(edad_gestacional_semanas__gte=33, edad_gestacional_semanas__lte=36).count()),
    ]
    
    for label, count in prematuros_data:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = count
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # 4. Contacto Piel a Piel
    row += 1
    ws.cell(row=row, column=1).value = "CONTACTO INMEDIATO PIEL A PIEL"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    apego_data = [
        ('RN <2.5kg con apego', partos.filter(recien_nacidos__peso__lt=2.5, recien_nacidos__apego_piel_a_piel=True).count()),
        ('RN ≥2.5kg con apego', partos.filter(recien_nacidos__peso__gte=2.5, recien_nacidos__apego_piel_a_piel=True).count()),
    ]
    
    for label, count in apego_data:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = count
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # 5. Lactancia Materna Precoz
    row += 1
    ws.cell(row=row, column=1).value = "LACTANCIA MATERNA ANTES DE 1 HORA"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    lactancia_data = [
        ('RN <2.5kg con lactancia', partos.filter(recien_nacidos__peso__lt=2.5, recien_nacidos__lactancia_precoz=True).count()),
        ('RN ≥2.5kg con lactancia', partos.filter(recien_nacidos__peso__gte=2.5, recien_nacidos__lactancia_precoz=True).count()),
    ]
    
    for label, count in lactancia_data:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = count
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # 6. Nacionalidad de la Madre
    row += 1
    ws.cell(row=row, column=1).value = "NACIONALIDAD DE LA MADRE"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    madres_chilenas = partos.filter(madre__nacionalidad='chilena').count()
    madres_extranjeras = partos.filter(madre__nacionalidad__isnull=False).exclude(madre__nacionalidad='chilena').count()
    
    ws.cell(row=row, column=1).value = "Madre Chilena"
    ws.cell(row=row, column=2).value = madres_chilenas
    for col in range(1, 3):
        ws.cell(row=row, column=col).border = border
    row += 1
    
    ws.cell(row=row, column=1).value = "Madre Extranjera"
    ws.cell(row=row, column=2).value = madres_extranjeras
    for col in range(1, 3):
        ws.cell(row=row, column=col).border = border
    row += 1
    
    # 7. Acompañante
    row += 1
    ws.cell(row=row, column=1).value = "CON ACOMPAÑANTE"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    con_acompanante = partos.filter(acompanante__isnull=False).exclude(acompanante='').count()
    ws.cell(row=row, column=1).value = "Partos con Acompañante"
    ws.cell(row=row, column=2).value = con_acompanante
    for col in range(1, 3):
        ws.cell(row=row, column=col).border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15
    
    return wb


def crear_libro_excel_seccion_b(fecha_inicio, fecha_fin):
    """
    Exporta Sección B: Interrupción del Embarazo (Abortos)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sección B - Abortos"
    
    # Estilos
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:D1')
    titulo = ws['A1']
    titulo.value = f"SECCIÓN B: INTERRUPCIÓN DEL EMBARAZO"
    titulo.font = Font(bold=True, size=14)
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:D2')
    subtitulo = ws['A2']
    subtitulo.value = f"Período: {fecha_inicio} a {fecha_fin}"
    subtitulo.font = Font(size=10, italic=True)
    subtitulo.alignment = Alignment(horizontal='center')
    
    # Obtener datos
    abortos = Aborto.objects.filter(fecha_derivacion__date__range=[fecha_inicio, fecha_fin])
    total_abortos = abortos.count()
    
    # Fila 4: Total
    row = 4
    ws.cell(row=row, column=1).value = "TOTAL DE ABORTOS"
    ws.cell(row=row, column=2).value = total_abortos
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF")
    row += 2
    
    # Encabezados
    headers = ['Tipo', 'Cantidad']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    row += 1
    
    # Tipos de Aborto
    tipos_abortos = abortos.values('tipo').annotate(count=Count('id')).order_by('tipo')
    for tipo_aborto in tipos_abortos:
        tipo = tipo_aborto['tipo'] or 'Sin especificar'
        count = tipo_aborto['count']
        
        ws.cell(row=row, column=1).value = tipo
        ws.cell(row=row, column=2).value = count
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        row += 1
    
    # Causales de Aborto
    row += 2
    ws.cell(row=row, column=1).value = "CAUSALES"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    causales = abortos.values('causal').annotate(count=Count('id')).order_by('causal')
    for causal_aborto in causales:
        causal = causal_aborto['causal'] or 'Sin especificar'
        count = causal_aborto['count']
        
        ws.cell(row=row, column=1).value = causal
        ws.cell(row=row, column=2).value = count
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # Ajustar ancho
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    
    return wb


def crear_libro_excel_seccion_c(fecha_inicio, fecha_fin):
    """
    Exporta Sección C: Métodos Anticonceptivos al Alta
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sección C - Anticoncepción"
    
    # Estilos
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:C1')
    titulo = ws['A1']
    titulo.value = f"SECCIÓN C: MÉTODOS ANTICONCEPTIVOS AL ALTA"
    titulo.font = Font(bold=True, size=14)
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:C2')
    subtitulo = ws['A2']
    subtitulo.value = f"Período: {fecha_inicio} a {fecha_fin}"
    subtitulo.font = Font(size=10, italic=True)
    subtitulo.alignment = Alignment(horizontal='center')
    
    # Obtener datos
    altas = Alta.objects.filter(madre__isnull=False, fecha_creacion__date__range=[fecha_inicio, fecha_fin])
    total_altas = altas.count()
    altas_con_anticonceptivo = altas.filter(se_entrego_anticonceptivo=True).count()
    
    # Fila 4: Resumen
    row = 4
    ws.cell(row=row, column=1).value = "Total de Altas de Madres"
    ws.cell(row=row, column=2).value = total_altas
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    ws.cell(row=row, column=1).value = "Altas con Anticonceptivo"
    ws.cell(row=row, column=2).value = altas_con_anticonceptivo
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    porcentaje_cobertura = round((altas_con_anticonceptivo / total_altas * 100) if total_altas > 0 else 0, 1)
    ws.cell(row=row, column=1).value = "Cobertura (%)"
    ws.cell(row=row, column=2).value = porcentaje_cobertura
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 2
    
    # Métodos entregados
    ws.cell(row=row, column=1).value = "MÉTODOS ENTREGADOS"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    headers = ['Método', 'Cantidad', 'Porcentaje']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    row += 1
    
    metodos = altas.filter(se_entrego_anticonceptivo=True).values('metodo_anticonceptivo').annotate(count=Count('id'))
    for metodo in metodos:
        metodo_name = metodo['metodo_anticonceptivo'] or 'Sin especificar'
        count = metodo['count']
        porcentaje = round((count / altas_con_anticonceptivo * 100) if altas_con_anticonceptivo > 0 else 0, 1)
        
        ws.cell(row=row, column=1).value = metodo_name
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = f"{porcentaje}%"
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        row += 1
    
    # Ajustar ancho
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    return wb


def crear_libro_excel_seccion_d(fecha_inicio, fecha_fin):
    """
    Exporta Sección D: Información de Recién Nacidos
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sección D - RN"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:D1')
    titulo = ws['A1']
    titulo.value = f"SECCIÓN D: INFORMACIÓN DE RECIÉN NACIDOS"
    titulo.font = Font(bold=True, size=14)
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:D2')
    subtitulo = ws['A2']
    subtitulo.value = f"Período: {fecha_inicio} a {fecha_fin}"
    subtitulo.font = Font(size=10, italic=True)
    subtitulo.alignment = Alignment(horizontal='center')
    
    # Obtener datos
    rn = RecienNacido.objects.filter(fecha_registro__date__range=[fecha_inicio, fecha_fin])
    total_rn = rn.count()
    
    # Fila 4: Totales generales
    row = 4
    ws.cell(row=row, column=1).value = "Total de RN"
    ws.cell(row=row, column=2).value = total_rn
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 2
    
    # Sexo
    ws.cell(row=row, column=1).value = "POR SEXO"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    sexo_data = [
        ('Masculino', rn.filter(sexo='M').count()),
        ('Femenino', rn.filter(sexo='F').count()),
    ]
    
    for label, count in sexo_data:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = count
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # APGAR
    row += 1
    ws.cell(row=row, column=1).value = "APGAR (1 minuto)"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    apgar_data = [
        ('APGAR < 7', rn.filter(apgar_1_min__lt=7).count()),
        ('APGAR ≥ 7', rn.filter(apgar_1_min__gte=7).count()),
    ]
    
    for label, count in apgar_data:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = count
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # Peso
    row += 1
    ws.cell(row=row, column=1).value = "POR PESO"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    peso_data = [
        ('Bajo Peso (<2.5kg)', rn.filter(peso__lt=2.5).count()),
        ('Peso Normal (2.5-3.9kg)', rn.filter(peso__gte=2.5, peso__lte=3.9).count()),
        ('Macrosomía (≥4.0kg)', rn.filter(peso__gte=4.0).count()),
    ]
    
    for label, count in peso_data:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = count
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # Talla
    row += 1
    ws.cell(row=row, column=1).value = "POR TALLA"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    promedio_talla = rn.values_list('talla', flat=True)
    promedio_talla_val = sum(filter(None, promedio_talla)) / len([x for x in promedio_talla if x]) if promedio_talla.count() > 0 else 0
    
    ws.cell(row=row, column=1).value = "Talla Promedio (cm)"
    ws.cell(row=row, column=2).value = round(promedio_talla_val, 2) if promedio_talla_val else 0
    for col in range(1, 3):
        ws.cell(row=row, column=col).border = border
    row += 1
    
    # Lactancia
    row += 1
    ws.cell(row=row, column=1).value = "LACTANCIA PRECOZ"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    lactancia_precoz = rn.filter(lactancia_precoz=True).count()
    ws.cell(row=row, column=1).value = "RN con Lactancia Precoz"
    ws.cell(row=row, column=2).value = lactancia_precoz
    for col in range(1, 3):
        ws.cell(row=row, column=col).border = border
    row += 1
    
    # Apego Piel a Piel
    row += 1
    ws.cell(row=row, column=1).value = "APEGO PIEL A PIEL"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    apego = rn.filter(apego_piel_a_piel=True).count()
    ws.cell(row=row, column=1).value = "RN con Apego Piel a Piel"
    ws.cell(row=row, column=2).value = apego
    for col in range(1, 3):
        ws.cell(row=row, column=col).border = border
    
    # Ajustar ancho
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    
    return wb


def descargar_excel_response(wb, nombre_archivo):
    """
    Genera una respuesta HTTP para descargar un archivo Excel.
    """
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response
